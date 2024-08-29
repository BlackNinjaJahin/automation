from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
from datetime import datetime


# Function to get the longest and shortest suggestions
def get_suggestions(search_term):
    driver.get("https://www.google.com")
    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys(search_term)

    # Wait for suggestions to load
    driver.implicitly_wait(20)

    suggestions = driver.find_elements(By.CSS_SELECTOR, 'ul[role="listbox"] li span')
    suggestions_text = [suggestion.text for suggestion in suggestions if suggestion.text]

    if suggestions_text:
        longest = max(suggestions_text, key=len)
        shortest = min(suggestions_text, key=len)
        return longest, shortest
    return None, None

# Load the Excel file
file_path = 'D:\\4BeatsQ1.xlsx'  # Replace with your actual file path
workbook = openpyxl.load_workbook(file_path)

# Determine today's day of the week
today_day = datetime.now().strftime('%A')  # 'Monday', 'Tuesday', ...

# Print available sheet names for debugging
print("Available sheets:", workbook.sheetnames)

# Try to select the sheet based on the current day
try:
    sheet = workbook[today_day]
except KeyError:
    print(f"Sheet for {today_day} does not exist.")
    # Fallback to a default sheet if needed
    sheet = workbook[workbook.sheetnames[0]]  # Fallback to the first sheet, or specify another sheet

# Initialize Chrome WebDriver
driver = webdriver.Chrome()  # Ensure ChromeDriver is installed and in PATH

# Iterate through each row in the selected sheet
for row in range(2, sheet.max_row + 1):  # Assuming the first row is the header
    keyword = sheet.cell(row=row, column=1).value  # Assuming keywords are in the first column

    if keyword:
        longest_option, shortest_option = get_suggestions(keyword)

        # Write results back to Excel
        sheet.cell(row=row, column=2).value = longest_option  # Assuming column 2 for longest
        sheet.cell(row=row, column=3).value = shortest_option  # Assuming column 3 for shortest


# Save the updated Excel file
workbook.save(file_path)
driver.quit()
